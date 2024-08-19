import os
import pandas as pd
from openpyxl import load_workbook
from flask import Flask, render_template, request, redirect, url_for

app = Flask(__name__)

def append_to_excel(file_path, data):
    backup_file_path = 'status_updates_backup.xlsx'
    
    df_new = pd.DataFrame(data, columns=['Name', 'Status', 'Date', 'Time'])

    try:
        if os.path.isfile(file_path):
            try:
                book = load_workbook(file_path)
                sheet = book['StatusUpdates']

                next_row = sheet.max_row + 1

                for idx, row in df_new.iterrows():
                    for col_idx, value in enumerate(row, start=1):
                        cell = sheet.cell(row=next_row, column=col_idx)
                        cell.value = value

                    next_row += 1

                book.save(file_path)
            
            except KeyError:
                df_existing = pd.DataFrame(columns=['Name', 'Status', 'Date', 'Time'])
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)

                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df_combined.to_excel(writer, sheet_name='StatusUpdates', index=False)

        else:
            df_new.to_excel(file_path, sheet_name='StatusUpdates', index=False)

    except PermissionError:
        with pd.ExcelWriter(backup_file_path, engine='openpyxl') as writer:
            df_new.to_excel(writer, sheet_name='StatusUpdates', index=False)
        print(f"Permission denied. Data written to backup file: {backup_file_path}")

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        name = request.form['name']
        status = request.form['status']
        date = request.form['date']
        time = request.form['time']

        data = [[name, status, date, time]]
        file_path = 'status_updates.xlsx'

        append_to_excel(file_path, data)

        return render_template('form.html', show_greeting=True)

    return render_template('form.html', show_greeting=False)

if __name__ == '__main__':
    app.run(debug=True)
